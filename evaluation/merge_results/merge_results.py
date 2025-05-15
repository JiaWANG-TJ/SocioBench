#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
结果合并工具
从各个模型的结果文件中提取指标并生成Excel比较表
"""

import os
import json
import glob
import pandas as pd
import re
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple, Union
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

# 指定输入和输出目录（使用绝对路径）
# 获取当前脚本的绝对路径
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
# 基于脚本位置计算结果目录
RESULTS_DIR = os.path.abspath(os.path.join(SCRIPT_DIR, "..", "results"))
OUTPUT_DIR = SCRIPT_DIR

# 需要提取的指标列表
METRICS = ["accuracy", "macro_f1", "micro_f1", "option_distance"]

# 百分比格式指标列表（这些指标将以百分比格式显示，保留两位小数）
PERCENTAGE_METRICS = ["accuracy", "macro_f1", "micro_f1"]

def parse_filename(filename: str) -> Tuple[str, str]:
    """
    从文件名解析模型名称和领域名称
    
    Args:
        filename: 结果文件的完整路径
        
    Returns:
        model_name: 模型名称
        domain: 领域名称
    """
    # 从路径中获取文件名
    base_name = os.path.basename(filename)
    
    # 解析文件名以获取领域和模型名称
    parts = base_name.split("__")
    if len(parts) >= 1:
        domain = parts[0]
    else:
        domain = "unknown"
    
    # 从路径中获取模型名称
    model_path = os.path.dirname(filename)
    model_name = os.path.basename(model_path)
    
    return model_name, domain

def extract_model_size(model_name: str) -> float:
    """
    从模型名称中提取参数大小（以十亿为单位）
    
    Args:
        model_name: 模型名称，如 'Qwen2.5-32B-Instruct' 或 'glm-4-9b-chat'
        
    Returns:
        size: 模型参数大小，单位为十亿（float类型）。如果无法提取，返回0.0
    """
    # 尝试匹配常见的参数大小模式
    # 匹配数字后跟B, b, Billion等标识
    patterns = [
        r'(\d+\.?\d*)B', # 匹配32B, 3.5B等格式
        r'(\d+\.?\d*)b', # 匹配9b, 7b等格式
        r'(\d+\.?\d*)[Bb]illion', # 匹配7Billion等格式
        r'-(\d+\.?\d*)[Bb]', # 匹配-32B, -7b等格式
    ]
    
    for pattern in patterns:
        match = re.search(pattern, model_name)
        if match:
            try:
                return float(match.group(1))
            except ValueError:
                pass
    
    # 如果没有找到匹配，尝试其他模式
    # 某些模型可能只有一个数字标识
    if re.search(r'[^\d](\d+)[^\d]', model_name):
        match = re.search(r'[^\d](\d+)[^\d]', model_name)
        try:
            # 如果数字很大（如大于100），可能是百万级别
            size = float(match.group(1))
            if size > 100:
                return size / 1000  # 转换为十亿
            return size
        except (ValueError, IndexError):
            pass
    
    # 如果无法提取，返回默认值
    print(f"无法从模型名称 '{model_name}' 中提取参数大小，将使用默认顺序")
    return 0.0

def sort_dataframe_by_model_size(df: pd.DataFrame) -> pd.DataFrame:
    """
    按照模型参数大小排序DataFrame的行
    
    Args:
        df: 包含模型性能指标的DataFrame
        
    Returns:
        sorted_df: 排序后的DataFrame
    """
    if df.empty:
        return df
    
    # 创建模型大小字典
    model_sizes = {}
    for model_name in df.index:
        model_sizes[model_name] = extract_model_size(model_name)
    
    # 创建一个包含模型大小的临时Series
    size_series = pd.Series(model_sizes)
    
    # 按模型大小排序DataFrame
    sorted_df = df.copy()
    sorted_df['_model_size'] = size_series
    sorted_df = sorted_df.sort_values('_model_size')
    sorted_df = sorted_df.drop('_model_size', axis=1)
    
    return sorted_df

def extract_metrics(file_path: str) -> Dict[str, float]:
    """
    从结果文件中提取指标
    
    Args:
        file_path: 结果文件的路径
        
    Returns:
        metrics_dict: 包含提取的指标的字典
    """
    metrics_dict = {}
    
    try:
        # 读取JSON文件的前几行来提取指标
        with open(file_path, 'r', encoding='utf-8') as f:
            # 只读取文件的前1000行，因为指标通常在文件开头
            content = ""
            for i, line in enumerate(f):
                content += line
                if i >= 1000:
                    break
            
            # 尝试解析JSON
            try:
                data = json.loads(content)
                # 提取需要的指标
                for metric in METRICS:
                    if metric in data:
                        metrics_dict[metric] = data[metric]
            except json.JSONDecodeError:
                # 如果不是有效的JSON，尝试查找特定的模式
                for metric in METRICS:
                    pattern = f'"{metric}":'
                    if pattern in content:
                        start_idx = content.find(pattern) + len(pattern)
                        end_idx = content.find(",", start_idx)
                        if end_idx == -1:  # 如果没有找到逗号，可能是最后一个元素
                            end_idx = content.find("}", start_idx)
                        if end_idx != -1:
                            value_str = content[start_idx:end_idx].strip()
                            try:
                                metrics_dict[metric] = float(value_str)
                            except ValueError:
                                pass
    except Exception as e:
        print(f"处理文件 {file_path} 出错: {e}")
    
    return metrics_dict

def process_all_results() -> Dict[str, pd.DataFrame]:
    """
    处理所有结果文件并为每个指标创建DataFrame
    
    Returns:
        metrics_dfs: 包含每个指标的DataFrame的字典
    """
    # 为每个指标创建一个数据字典
    metrics_data = {metric: {} for metric in METRICS}
    
    # 跟踪所有模型和领域
    all_models = set()
    all_domains = set()
    
    # 打印当前使用的目录信息
    print(f"搜索结果文件的目录: {RESULTS_DIR}")
    print(f"输出目录: {OUTPUT_DIR}")
    
    # 首先检查目录是否存在
    if not os.path.exists(RESULTS_DIR):
        print(f"错误: 结果目录不存在: {RESULTS_DIR}")
        # 尝试列出父目录的内容以帮助调试
        parent_dir = os.path.dirname(RESULTS_DIR)
        if os.path.exists(parent_dir):
            print(f"父目录 {parent_dir} 的内容:")
            try:
                for item in os.listdir(parent_dir):
                    print(f"  - {item}")
            except Exception as e:
                print(f"列出父目录内容时出错: {e}")
        return {metric: pd.DataFrame() for metric in METRICS}
    
    # 查找所有结果文件
    result_pattern = os.path.join(RESULTS_DIR, "**", "**", "*__results_*.json")
    print(f"使用文件匹配模式: {result_pattern}")
    
    result_files = glob.glob(result_pattern, recursive=True)
    
    print(f"找到 {len(result_files)} 个结果文件")
    
    # 如果没有找到文件，尝试列出目录内容以帮助调试
    if len(result_files) == 0:
        print("尝试列出结果目录结构以帮助调试:")
        for root, dirs, files in os.walk(RESULTS_DIR):
            level = root.replace(RESULTS_DIR, '').count(os.sep)
            indent = ' ' * 4 * level
            print(f"{indent}{os.path.basename(root)}/")
            sub_indent = ' ' * 4 * (level + 1)
            for file in files:
                print(f"{sub_indent}{file}")
        
        # 尝试使用不同的匹配模式
        alt_pattern = os.path.join(RESULTS_DIR, "**", "*results*.json")
        alt_files = glob.glob(alt_pattern, recursive=True)
        print(f"使用替代模式 '{alt_pattern}' 找到了 {len(alt_files)} 个文件")
        if len(alt_files) > 0:
            print("找到的文件:")
            for file in alt_files[:10]:  # 只显示前10个
                print(f"  - {file}")
            
            # 使用替代的文件列表
            result_files = alt_files
    
    # 处理每个结果文件
    for file_path in result_files:
        model_name, domain = parse_filename(file_path)
        metrics = extract_metrics(file_path)
        
        # 添加到跟踪集合
        all_models.add(model_name)
        all_domains.add(domain)
        
        # 将指标添加到相应的数据字典中
        for metric, value in metrics.items():
            if model_name not in metrics_data[metric]:
                metrics_data[metric][model_name] = {}
            metrics_data[metric][model_name][domain] = value
    
    # 将每个指标的数据转换为DataFrame并排序
    metrics_dfs = {}
    for metric in METRICS:
        # 创建从模型到领域的字典
        df_data = {}
        for model in all_models:
            if model in metrics_data[metric]:
                df_data[model] = metrics_data[metric][model]
        
        # 转换为DataFrame
        if df_data:
            df = pd.DataFrame.from_dict(df_data, orient='index')
            # 按照模型参数大小排序
            metrics_dfs[metric] = sort_dataframe_by_model_size(df)
        else:
            metrics_dfs[metric] = pd.DataFrame()
    
    return metrics_dfs

def apply_formatting(output_file: str, df: pd.DataFrame, metric: str) -> None:
    """
    对Excel文件应用格式，对每个领域最大值加粗，第二大值下划线,
    并为特定指标应用百分比格式
    
    Args:
        output_file: Excel文件路径
        df: 原始数据DataFrame
        metric: 指标名称
    """
    try:
        # 加载工作簿
        workbook = openpyxl.load_workbook(output_file)
        sheet = workbook.active
        
        # 计算每列的起始索引（考虑到索引列）
        col_offset = 2  # Excel中的列B对应于DataFrame的第一列
        
        # 为每一列（领域）找出最大值和第二大值
        for col_idx, col_name in enumerate(df.columns):
            # 获取该列的所有非空值
            col_values = df[col_name].dropna().tolist()
            
            if len(col_values) >= 2:  # 确保至少有两个值可以比较
                # 排序找出最大和第二大的值
                sorted_values = sorted(col_values, reverse=True)
                max_value = sorted_values[0]
                second_max_value = sorted_values[1]
                
                # 查找最大值和第二大值在DataFrame中的行索引
                max_idx = None
                second_max_idx = None
                
                for row_idx, value in enumerate(df[col_name]):
                    if pd.notna(value):
                        if value == max_value and max_idx is None:
                            max_idx = row_idx
                        elif value == second_max_value and second_max_idx is None:
                            second_max_idx = row_idx
                
                # 在Excel中应用格式（行索引+2是因为Excel是1开始并且有一个标题行）
                if max_idx is not None:
                    cell = sheet.cell(row=max_idx+2, column=col_idx+col_offset)
                    cell.font = Font(bold=True)
                
                if second_max_idx is not None:
                    cell = sheet.cell(row=second_max_idx+2, column=col_idx+col_offset)
                    cell.font = Font(underline='single')
        
        # 如果是百分比格式指标，设置数字格式
        if metric in PERCENTAGE_METRICS:
            # 数值列的起始索引
            for col_idx in range(len(df.columns)):
                # 获取列的字母标识
                col_letter = get_column_letter(col_idx + col_offset)
                
                # 格式化此列中的所有单元格
                for row_idx in range(2, len(df) + 2):  # 从第2行开始（跳过标题）
                    cell = sheet[f"{col_letter}{row_idx}"]
                    
                    # 应用百分比格式（不显示百分号，保留两位小数）
                    cell.number_format = '0.00'
                    
                    # 如果单元格有值，将数值乘以100
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        cell.value = cell.value * 100
                        
        # 保存修改后的工作簿
        workbook.save(output_file)
        print(f"已应用格式化到 {output_file}")
    except Exception as e:
        print(f"应用格式化到 {output_file} 时出错: {e}")

def save_to_excel(metrics_dfs: Dict[str, pd.DataFrame]) -> None:
    """
    将每个指标的DataFrame保存为Excel文件，并对最大值和第二大值应用格式
    
    Args:
        metrics_dfs: 包含每个指标的DataFrame的字典
    """
    # 保存每个指标的DataFrame
    for metric, df in metrics_dfs.items():
        if not df.empty:
            # 创建输出文件路径（不使用时间戳）
            output_file = os.path.join(OUTPUT_DIR, f"{metric}_comparison.xlsx")
            
            # 保存到Excel
            df.to_excel(output_file)
            print(f"保存 {metric} 比较结果到 {output_file}")
            
            # 应用格式化
            apply_formatting(output_file, df, metric)
        else:
            print(f"没有 {metric} 的数据可以保存")

def main():
    """主函数，处理所有结果并生成Excel文件"""
    print("开始处理结果...")
    
    # 处理所有结果文件
    metrics_dfs = process_all_results()
    
    # 检查是否有数据
    has_data = any(not df.empty for df in metrics_dfs.values())
    if not has_data:
        print("警告: 没有找到可以处理的数据，无法生成Excel文件")
    else:
        # 保存结果到Excel
        save_to_excel(metrics_dfs)
    
    print("处理完成！")

if __name__ == "__main__":
    main() 